"""
Report Builder Routes - Intelligent Report Generation System
Features:
- Sales reports (pillar-wise and consolidated)
- Service ticket analytics
- CSV/Excel export
- Email scheduling
"""

import io
import csv
from datetime import datetime, timezone, timedelta
from typing import Optional
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
import logging

logger = logging.getLogger(__name__)

# Create router
report_builder_router = APIRouter(prefix="/api/admin/reports", tags=["Report Builder"])

# Database reference (set during startup)
db = None

def set_report_builder_db(database):
    global db
    db = database

# Dependency placeholder for admin verification
verify_admin = None

def set_verify_admin(verify_fn):
    global verify_admin
    verify_admin = verify_fn


def get_admin_verifier():
    """Wrapper to properly invoke verify_admin"""
    if verify_admin:
        return verify_admin
    raise HTTPException(status_code=401, detail="Not authenticated")


@report_builder_router.get("/generate")
async def generate_report(
    report_type: str = "daily_summary",
    period: str = "today",
    pillar: str = "all",
    start_date: str = None,
    end_date: str = None,
    username: str = Depends(get_admin_verifier)
):
    """Generate intelligent report based on type and parameters"""
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif period == "yesterday":
        start = (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = start.replace(hour=23, minute=59, second=59)
    elif period == "this_week":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif period == "last_week":
        start = (now - timedelta(days=now.weekday() + 7)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=6, hours=23, minutes=59, seconds=59)
    elif period == "this_month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif period == "last_month":
        first_of_month = now.replace(day=1)
        last_month_end = first_of_month - timedelta(days=1)
        start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = last_month_end.replace(hour=23, minute=59, second=59)
    elif period == "last_30_days":
        start = now - timedelta(days=30)
        end = now
    elif period == "last_90_days":
        start = now - timedelta(days=90)
        end = now
    elif period == "this_year":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
    elif period == "custom" and start_date and end_date:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        start = now - timedelta(days=7)
        end = now
    
    date_filter = {"$gte": start.isoformat(), "$lte": end.isoformat()}
    pillar_filter = {"pillar": pillar} if pillar != "all" else {}
    
    # Generate report based on type
    if report_type == "daily_summary":
        # Orders
        orders = await db.orders.find({"created_at": date_filter}).to_list(10000)
        total_revenue = sum(o.get("total", 0) for o in orders)
        total_orders = len(orders)
        avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
        
        # Tickets
        tickets = await db.service_desk_tickets.find({"created_at": date_filter}).to_list(10000)
        open_tickets = len([t for t in tickets if t.get("status") in ["open", "pending"]])
        resolved_tickets = len([t for t in tickets if t.get("status") in ["resolved", "closed"]])
        
        # New Members
        users = await db.users.find({"created_at": date_filter}).to_list(10000)
        new_members = len(users)
        
        # Pets registered
        pets = await db.pets.find({"created_at": date_filter}).to_list(10000)
        new_pets = len(pets)
        
        return {
            "report_type": "daily_summary",
            "period": period,
            "date_range": {"start": start.isoformat(), "end": end.isoformat()},
            "summary": [
                {"label": "Total Revenue", "value": f"₹{total_revenue:,.0f}", "change": None},
                {"label": "Orders", "value": str(total_orders), "change": None},
                {"label": "Avg Order Value", "value": f"₹{avg_order_value:,.0f}", "change": None},
                {"label": "New Members", "value": str(new_members), "change": None},
                {"label": "New Pets", "value": str(new_pets), "change": None},
                {"label": "Open Tickets", "value": str(open_tickets), "change": None},
                {"label": "Resolved Tickets", "value": str(resolved_tickets), "change": None}
            ],
            "columns": ["Date", "Orders", "Revenue", "New Members", "Tickets"],
            "rows": []
        }
    
    elif report_type == "pillar_performance":
        # Get revenue by pillar
        orders = await db.orders.find({"created_at": date_filter}).to_list(10000)
        pillar_data = {}
        
        for order in orders:
            p = order.get("pillar", "shop")
            if p not in pillar_data:
                pillar_data[p] = {"orders": 0, "revenue": 0}
            pillar_data[p]["orders"] += 1
            pillar_data[p]["revenue"] += order.get("total", 0)
        
        rows = []
        for p, data in pillar_data.items():
            rows.append([p.capitalize(), str(data["orders"]), f"₹{data['revenue']:,.0f}"])
        
        total_revenue = sum(d["revenue"] for d in pillar_data.values())
        
        return {
            "report_type": "pillar_performance",
            "period": period,
            "summary": [
                {"label": "Total Revenue", "value": f"₹{total_revenue:,.0f}"},
                {"label": "Pillars Active", "value": str(len(pillar_data))},
            ],
            "columns": ["Pillar", "Orders", "Revenue"],
            "rows": rows
        }
    
    elif report_type == "order_report":
        orders = await db.orders.find({"created_at": date_filter, **pillar_filter}).to_list(10000)
        
        rows = []
        for o in orders[:100]:
            rows.append([
                o.get("order_id", o.get("id", "N/A")),
                o.get("created_at", "")[:10],
                o.get("customer_name", o.get("customer", {}).get("name", "N/A")),
                o.get("pillar", "shop"),
                f"₹{o.get('total', 0):,.0f}",
                o.get("status", "pending")
            ])
        
        return {
            "report_type": "order_report",
            "period": period,
            "summary": [
                {"label": "Total Orders", "value": str(len(orders))},
                {"label": "Total Revenue", "value": f"₹{sum(o.get('total', 0) for o in orders):,.0f}"}
            ],
            "columns": ["Order ID", "Date", "Customer", "Pillar", "Amount", "Status"],
            "rows": rows
        }
    
    elif report_type == "ticket_report":
        tickets = await db.service_desk_tickets.find({"created_at": date_filter}).to_list(10000)
        
        status_counts = {}
        pillar_counts = {}
        urgency_counts = {}
        
        for t in tickets:
            s = t.get("status", "open")
            p = t.get("pillar", "general")
            u = t.get("urgency", "normal")
            status_counts[s] = status_counts.get(s, 0) + 1
            pillar_counts[p] = pillar_counts.get(p, 0) + 1
            urgency_counts[u] = urgency_counts.get(u, 0) + 1
        
        rows = []
        for t in tickets[:100]:
            rows.append([
                t.get("ticket_id", "N/A"),
                t.get("created_at", "")[:10],
                t.get("subject", "No subject")[:50],
                t.get("pillar", "general"),
                t.get("status", "open"),
                t.get("urgency", "normal")
            ])
        
        return {
            "report_type": "ticket_report",
            "period": period,
            "summary": [
                {"label": "Total Tickets", "value": str(len(tickets))},
                {"label": "Open", "value": str(status_counts.get("open", 0) + status_counts.get("pending", 0))},
                {"label": "In Progress", "value": str(status_counts.get("in_progress", 0))},
                {"label": "Resolved", "value": str(status_counts.get("resolved", 0) + status_counts.get("closed", 0))},
                {"label": "Critical", "value": str(urgency_counts.get("critical", 0))},
                {"label": "High Priority", "value": str(urgency_counts.get("high", 0))}
            ],
            "columns": ["Ticket ID", "Date", "Subject", "Pillar", "Status", "Priority"],
            "rows": rows,
            "analytics": {
                "by_status": status_counts,
                "by_pillar": pillar_counts,
                "by_urgency": urgency_counts
            }
        }
    
    elif report_type == "revenue_report":
        orders = await db.orders.find({"created_at": date_filter, **pillar_filter}).to_list(10000)
        
        # Calculate GST breakdown
        total_subtotal = 0
        total_gst = 0
        total_shipping = 0
        total_discounts = 0
        
        rows = []
        for o in orders:
            pricing = o.get("pricing", {})
            subtotal = pricing.get("subtotal", o.get("subtotal", o.get("total", 0)))
            gst = pricing.get("gst_details", {}).get("total_tax", 0)
            shipping = pricing.get("shipping_fee", o.get("shipping_fee", 0))
            discount = pricing.get("discount_amount", o.get("discount_amount", 0))
            
            total_subtotal += subtotal
            total_gst += gst
            total_shipping += shipping
            total_discounts += discount
            
            rows.append([
                o.get("order_id", "N/A"),
                o.get("created_at", "")[:10],
                f"₹{subtotal:,.0f}",
                f"₹{gst:,.0f}",
                f"₹{shipping:,.0f}",
                f"₹{discount:,.0f}",
                f"₹{o.get('total', subtotal + gst + shipping - discount):,.0f}"
            ])
        
        return {
            "report_type": "revenue_report",
            "period": period,
            "summary": [
                {"label": "Total Revenue", "value": f"₹{total_subtotal + total_gst + total_shipping - total_discounts:,.0f}"},
                {"label": "Subtotal", "value": f"₹{total_subtotal:,.0f}"},
                {"label": "GST Collected", "value": f"₹{total_gst:,.0f}"},
                {"label": "Shipping Revenue", "value": f"₹{total_shipping:,.0f}"},
                {"label": "Discounts Given", "value": f"₹{total_discounts:,.0f}"}
            ],
            "columns": ["Order ID", "Date", "Subtotal", "GST", "Shipping", "Discount", "Total"],
            "rows": rows[:100]
        }
    
    elif report_type == "member_analytics":
        users = await db.users.find({}).to_list(50000)
        
        # Handle both datetime and string created_at fields
        def get_created_at_str(u):
            created = u.get("created_at", "")
            if isinstance(created, datetime):
                return created.isoformat()
            return str(created) if created else ""
        
        start_iso = start.isoformat()
        new_in_period = [u for u in users if get_created_at_str(u) >= start_iso]
        
        tier_counts = {}
        for u in users:
            tier = u.get("membership_tier", "free")
            tier_counts[tier] = tier_counts.get(tier, 0) + 1
        
        return {
            "report_type": "member_analytics",
            "period": period,
            "summary": [
                {"label": "Total Members", "value": str(len(users))},
                {"label": "New This Period", "value": str(len(new_in_period))},
                {"label": "VIP Members", "value": str(tier_counts.get("vip", 0))},
                {"label": "Annual Members", "value": str(tier_counts.get("annual", 0))}
            ],
            "columns": ["Tier", "Count", "Percentage"],
            "rows": [[tier, str(count), f"{count/len(users)*100:.1f}%"] for tier, count in tier_counts.items()]
        }
    
    elif report_type == "pet_analytics":
        pets = await db.pets.find({}).to_list(50000)
        
        # Handle both datetime and string created_at fields
        def get_created_at_str(p):
            created = p.get("created_at", "")
            if isinstance(created, datetime):
                return created.isoformat()
            return str(created) if created else ""
        
        start_iso = start.isoformat()
        new_in_period = [p for p in pets if get_created_at_str(p) >= start_iso]
        
        scored = [p for p in pets if p.get("overall_score")]
        avg_score = sum(p.get("overall_score", 0) for p in scored) / len(scored) if scored else 0
        
        def format_created_date(p):
            created = p.get("created_at", "")
            if isinstance(created, datetime):
                return created.strftime("%Y-%m-%d")
            return str(created)[:10] if created else "N/A"
        
        return {
            "report_type": "pet_analytics",
            "period": period,
            "summary": [
                {"label": "Total Pets", "value": str(len(pets))},
                {"label": "New This Period", "value": str(len(new_in_period))},
                {"label": "With Soul Score", "value": str(len(scored))},
                {"label": "Avg Soul Score", "value": f"{avg_score:.0f}%"}
            ],
            "columns": ["Pet Name", "Breed", "Age", "Soul Score", "Created"],
            "rows": [[p.get("name", "N/A"), p.get("breed", "N/A"), str(p.get("age", "N/A")), f"{p.get('overall_score', 0):.0f}%", format_created_date(p)] for p in new_in_period[:50]]
        }
    
    elif report_type == "product_performance":
        products = await db.products.find({}).to_list(1000)
        orders = await db.orders.find({"created_at": date_filter}).to_list(10000)
        
        product_sales = {}
        for order in orders:
            for item in order.get("items", []):
                pid = item.get("id", item.get("product_id"))
                if pid:
                    if pid not in product_sales:
                        product_sales[pid] = {"name": item.get("name", "Unknown"), "qty": 0, "revenue": 0}
                    product_sales[pid]["qty"] += item.get("quantity", 1)
                    product_sales[pid]["revenue"] += item.get("price", 0) * item.get("quantity", 1)
        
        sorted_products = sorted(product_sales.items(), key=lambda x: x[1]["revenue"], reverse=True)
        
        return {
            "report_type": "product_performance",
            "period": period,
            "summary": [
                {"label": "Products Sold", "value": str(sum(p["qty"] for p in product_sales.values()))},
                {"label": "Unique Products", "value": str(len(product_sales))},
                {"label": "Total Revenue", "value": f"₹{sum(p['revenue'] for p in product_sales.values()):,.0f}"}
            ],
            "columns": ["Product", "Units Sold", "Revenue"],
            "rows": [[data["name"], str(data["qty"]), f"₹{data['revenue']:,.0f}"] for pid, data in sorted_products[:20]]
        }
    
    return {"error": "Unknown report type", "available_types": ["daily_summary", "pillar_performance", "order_report", "ticket_report", "revenue_report", "member_analytics", "pet_analytics", "product_performance"]}


@report_builder_router.get("/export/csv")
async def export_report_csv(
    report_type: str = "daily_summary",
    period: str = "today",
    pillar: str = "all",
    username: str = Depends(lambda: verify_admin)
):
    """Export report as CSV file"""
    
    # Generate the report data
    report = await generate_report(report_type, period, pillar, None, None, username)
    
    if "error" in report:
        raise HTTPException(status_code=400, detail=report["error"])
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    if report.get("columns"):
        writer.writerow(report["columns"])
    
    # Write rows
    for row in report.get("rows", []):
        writer.writerow(row)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={report_type}_{period}.csv"}
    )


@report_builder_router.get("/export/excel")
async def export_report_excel(
    report_type: str = "daily_summary",
    period: str = "today",
    pillar: str = "all",
    username: str = Depends(lambda: verify_admin)
):
    """Export report as Excel file (xlsx)"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return await export_report_csv(report_type, period, pillar, username)
    
    # Generate the report data
    report = await generate_report(report_type, period, pillar, None, None, username)
    
    if "error" in report:
        raise HTTPException(status_code=400, detail=report["error"])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace("_", " ").title()
    
    # Style definitions
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write summary section
    if report.get("summary"):
        ws.append(["Report Summary"])
        ws['A1'].font = Font(bold=True, size=14)
        for item in report["summary"]:
            ws.append([item["label"], item["value"]])
        ws.append([])
    
    # Write header
    start_row = ws.max_row + 1
    if report.get("columns"):
        ws.append(report["columns"])
        for col_idx, _ in enumerate(report["columns"], 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
    
    # Write rows
    for row in report.get("rows", []):
        ws.append(row)
        current_row = ws.max_row
        for col_idx in range(1, len(row) + 1):
            ws.cell(row=current_row, column=col_idx).border = border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return StreamingResponse(
        iter([excel_buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_type}_{period}.xlsx"}
    )


@report_builder_router.post("/schedule")
async def save_report_schedule(
    schedule: dict,
    username: str = Depends(lambda: verify_admin)
):
    """Save email report schedule configuration"""
    schedule["updated_at"] = datetime.now(timezone.utc).isoformat()
    schedule["updated_by"] = username
    
    await db.report_schedules.update_one(
        {"type": "email_schedule"},
        {"$set": schedule},
        upsert=True
    )
    
    return {"success": True, "message": "Schedule saved"}
